import json
import os
from typing import Dict, Any, List
from reportlab.lib.pagesizes import letter, A4
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.lib import colors
from io import BytesIO
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.chart import BarChart, Reference
import datetime

class ReportGenerator:
    def __init__(self):
        self.styles = getSampleStyleSheet()
        self.title_style = ParagraphStyle(
            'CustomTitle',
            parent=self.styles['Heading1'],
            fontSize=18,
            spaceAfter=30,
            textColor=colors.darkblue
        )
        self.heading_style = ParagraphStyle(
            'CustomHeading',
            parent=self.styles['Heading2'],
            fontSize=14,
            spaceAfter=12,
            textColor=colors.darkblue
        )
    
    def generate_pdf_report(self, business_data: Dict[str, Any], recommendations: Dict[str, Any]) -> BytesIO:
        """Generate  PDF report"""
        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=letter, rightMargin=72, leftMargin=72, topMargin=72, bottomMargin=18)
        
        # Build story
        story = []
        
        # Title
        story.append(Paragraph("Financial Planning Report", self.title_style))
        story.append(Spacer(1, 12))
        
        # Business Information Section
        story.append(Paragraph("Business Information", self.heading_style))
        business_info_data = [
            ['Company Location', f"{business_data.get('location', {}).get('country', 'N/A')}"],
            ['Industry', business_data.get('industry', 'N/A').title()],
            ['Number of Employees', str(business_data.get('employeeCount', 0))],
            ['Monthly Revenue', f"${business_data.get('monthlyRevenue', 0):,.2f}"],
            ['Currency', business_data.get('currency', 'USD')]
        ]
        
        business_table = Table(business_info_data, colWidths=[2*inch, 3*inch])
        business_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 12),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        story.append(business_table)
        story.append(Spacer(1, 20))
        
        # Emergency Fund Section
        story.append(Paragraph("Emergency Fund Recommendations", self.heading_style))
        emergency_fund = recommendations.get('emergencyFund', {})
        emergency_data = [
            ['Recommended Amount', f"${emergency_fund.get('recommendedAmount', 0):,.2f}"],
            ['Current Gap', f"${emergency_fund.get('currentGap', 0):,.2f}"],
            ['Monthly Contribution Needed', f"${emergency_fund.get('monthlyContribution', 0):,.2f}"],
            ['Time to Goal (months)', f"{emergency_fund.get('timeToGoal', 0):.1f}"]
        ]
        
        emergency_table = Table(emergency_data, colWidths=[2.5*inch, 2.5*inch])
        emergency_table.setStyle(self._get_table_style())
        story.append(emergency_table)
        story.append(Spacer(1, 20))
        
        # Growth Fund Section
        story.append(Paragraph("Growth Investment Recommendations", self.heading_style))
        growth_fund = recommendations.get('growthFund', {})
        growth_data = [
            ['Total Growth Budget', f"${growth_fund.get('totalBudget', 0):,.2f}"],
            ['Marketing Budget', f"${growth_fund.get('marketingBudget', 0):,.2f}"],
            ['Hiring Budget', f"${growth_fund.get('hiringBudget', 0):,.2f}"],
            ['Equipment Upgrade', f"${growth_fund.get('equipmentUpgrade', 0):,.2f}"],
            ['Research & Development', f"${growth_fund.get('researchDevelopment', 0):,.2f}"]
        ]
        
        growth_table = Table(growth_data, colWidths=[2.5*inch, 2.5*inch])
        growth_table.setStyle(self._get_table_style())
        story.append(growth_table)
        story.append(Spacer(1, 20))
        
        # Tax Planning Section
        story.append(Paragraph("Tax Planning Analysis", self.heading_style))
        tax_planning = recommendations.get('taxPlanning', {})
        tax_data = [
            ['Estimated Tax Liability', f"${tax_planning.get('estimatedTaxLiability', 0):,.2f}"],
            ['Corporate Tax', f"${tax_planning.get('corporateTax', 0):,.2f}"],
            ['Payroll Tax', f"${tax_planning.get('payrollTax', 0):,.2f}"],
            ['Annual Profit', f"${tax_planning.get('annualProfit', 0):,.2f}"]
        ]
        
        tax_table = Table(tax_data, colWidths=[2.5*inch, 2.5*inch])
        tax_table.setStyle(self._get_table_style())
        story.append(tax_table)
        story.append(Spacer(1, 20))
        
        # Retirement Planning Section
        story.append(Paragraph("Retirement Planning Recommendations", self.heading_style))
        retirement = recommendations.get('retirementPlanning', {})
        retirement_data = [
            ['Recommended Plan', retirement.get('recommendedPlan', 'N/A')],
            ['Maximum Contribution', f"${retirement.get('maxContribution', 0):,.2f}"],
            ['Recommended Contribution', f"${retirement.get('recommendedContribution', 0):,.2f}"],
            ['Tax Benefits', retirement.get('taxBenefits', 'N/A')]
        ]
        
        retirement_table = Table(retirement_data, colWidths=[2.5*inch, 2.5*inch])
        retirement_table.setStyle(self._get_table_style())
        story.append(retirement_table)
        story.append(Spacer(1, 20))
        
        # Cash Flow Forecast Summary
        story.append(Paragraph("Annual Cash Flow Summary", self.heading_style))
        cash_flow = recommendations.get('cashFlowForecast', {}).get('annualSummary', {})
        cash_flow_data = [
            ['Total Annual Revenue', f"${cash_flow.get('totalRevenue', 0):,.2f}"],
            ['Total Annual Expenses', f"${cash_flow.get('totalExpenses', 0):,.2f}"],
            ['Net Cash Flow', f"${cash_flow.get('totalNetCashFlow', 0):,.2f}"],
            ['Average Monthly Revenue', f"${cash_flow.get('averageMonthlyRevenue', 0):,.2f}"]
        ]
        
        cash_flow_table = Table(cash_flow_data, colWidths=[2.5*inch, 2.5*inch])
        cash_flow_table.setStyle(self._get_table_style())
        story.append(cash_flow_table)
        
        # Footer
        story.append(Spacer(1, 30))
        story.append(Paragraph(f"Report generated on {datetime.datetime.now().strftime('%B %d, %Y')}", self.styles['Normal']))
        story.append(Paragraph("This report is for informational purposes only and should not be considered as professional financial advice.", self.styles['Italic']))
        
        # Build PDF
        doc.build(story)
        buffer.seek(0)
        return buffer
    
    def generate_excel_report(self, business_data: Dict[str, Any], recommendations: Dict[str, Any]) -> BytesIO:
        """Generate  Excel report"""
        buffer = BytesIO()
        workbook = openpyxl.Workbook()
        
        # Remove default sheet
        workbook.remove(workbook.active)
        
        # Create Summary sheet
        summary_sheet = workbook.create_sheet("Summary")
        self._create_summary_sheet(summary_sheet, business_data, recommendations)
        
        # Create Cash Flow Forecast sheet
        forecast_sheet = workbook.create_sheet("Cash Flow Forecast")
        self._create_forecast_sheet(forecast_sheet, recommendations.get('cashFlowForecast', {}))
        
        # Create Recommendations sheet
        recommendations_sheet = workbook.create_sheet("Detailed Recommendations")
        self._create_recommendations_sheet(recommendations_sheet, recommendations)
        
        # Save to buffer
        workbook.save(buffer)
        buffer.seek(0)
        return buffer
    
    def _get_table_style(self):
        """Get standard table style"""
        return TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.lightblue),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.black),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.white),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ])
    
    def _create_summary_sheet(self, sheet, business_data: Dict[str, Any], recommendations: Dict[str, Any]):
        """Create summary sheet in Excel"""
        # Title
        sheet['A1'] = "Financial Planning Report Summary"
        sheet['A1'].font = Font(size=16, bold=True)
        sheet.merge_cells('A1:B1')
        
        # Business Information
        sheet['A3'] = "Business Information"
        sheet['A3'].font = Font(size=14, bold=True)
        
        row = 4
        business_info = [
            ('Location', business_data.get('location', {}).get('country', 'N/A')),
            ('Industry', business_data.get('industry', 'N/A').title()),
            ('Employees', business_data.get('employeeCount', 0)),
            ('Monthly Revenue', f"${business_data.get('monthlyRevenue', 0):,.2f}"),
            ('Currency', business_data.get('currency', 'USD'))
        ]
        
        for label, value in business_info:
            sheet[f'A{row}'] = label
            sheet[f'B{row}'] = value
            sheet[f'A{row}'].font = Font(bold=True)
            row += 1
        
        # Key Recommendations
        row += 2
        sheet[f'A{row}'] = "Key Recommendations"
        sheet[f'A{row}'].font = Font(size=14, bold=True)
        row += 1
        
        emergency_fund = recommendations.get('emergencyFund', {})
        growth_fund = recommendations.get('growthFund', {})
        
        key_recs = [
            ('Emergency Fund Needed', f"${emergency_fund.get('recommendedAmount', 0):,.2f}"),
            ('Monthly Emergency Contribution', f"${emergency_fund.get('monthlyContribution', 0):,.2f}"),
            ('Growth Investment Budget', f"${growth_fund.get('totalBudget', 0):,.2f}"),
            ('Estimated Annual Tax', f"${recommendations.get('taxPlanning', {}).get('estimatedTaxLiability', 0):,.2f}")
        ]
        
        for label, value in key_recs:
            sheet[f'A{row}'] = label
            sheet[f'B{row}'] = value
            sheet[f'A{row}'].font = Font(bold=True)
            row += 1
        
        # Auto-adjust column widths
        for column in sheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except Exception:
                    pass
            adjusted_width = min(max_length + 2, 50)
            sheet.column_dimensions[column_letter].width = adjusted_width
    
    def _create_forecast_sheet(self, sheet, cash_flow_data: Dict[str, Any]):
        """Create cash flow forecast sheet"""
        # Title
        sheet['A1'] = "12-Month Cash Flow Forecast"
        sheet['A1'].font = Font(size=16, bold=True)
        sheet.merge_cells('A1:D1')
        
        # Headers
        headers = ['Month', 'Revenue', 'Expenses', 'Net Cash Flow']
        for col, header in enumerate(headers, 1):
            cell = sheet.cell(row=3, column=col)
            cell.value = header
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        
        # Data
        monthly_forecast = cash_flow_data.get('monthlyForecast', [])
        for row, month_data in enumerate(monthly_forecast, 4):
            sheet.cell(row=row, column=1).value = f"Month {month_data.get('month', 0)}"
            sheet.cell(row=row, column=2).value = month_data.get('revenue', 0)
            sheet.cell(row=row, column=3).value = month_data.get('expenses', 0)
            sheet.cell(row=row, column=4).value = month_data.get('netCashFlow', 0)
        
        # Format currency columns
        for row in range(4, 16):
            for col in range(2, 5):
                cell = sheet.cell(row=row, column=col)
                cell.number_format = '"$"#,##0.00'
        
        # Create chart
        chart = BarChart()
        chart.type = "col"
        chart.style = 10
        chart.title = "Monthly Cash Flow"
        chart.y_axis.title = 'Amount ($)'
        chart.x_axis.title = 'Month'
        
        data = Reference(sheet, min_col=2, min_row=3, max_row=15, max_col=4)
        cats = Reference(sheet, min_col=1, min_row=4, max_row=15)
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        
        sheet.add_chart(chart, "F3")
        
        # Auto-adjust column widths
        for column in sheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 20)
            sheet.column_dimensions[column_letter].width = adjusted_width
    
    def _create_recommendations_sheet(self, sheet, recommendations: Dict[str, Any]):
        """Create detailed recommendations sheet"""
        # Title
        sheet['A1'] = "Detailed Financial Recommendations"
        sheet['A1'].font = Font(size=16, bold=True)
        sheet.merge_cells('A1:B1')
        
        row = 3
        
        # Emergency Fund
        sheet[f'A{row}'] = "Emergency Fund"
        sheet[f'A{row}'].font = Font(size=14, bold=True)
        row += 1
        
        emergency_fund = recommendations.get('emergencyFund', {})
        emergency_items = [
            ('Recommended Amount', f"${emergency_fund.get('recommendedAmount', 0):,.2f}"),
            ('Current Gap', f"${emergency_fund.get('currentGap', 0):,.2f}"),
            ('Monthly Contribution', f"${emergency_fund.get('monthlyContribution', 0):,.2f}"),
            ('Time to Goal (months)', f"{emergency_fund.get('timeToGoal', 0):.1f}")
        ]
        
        for label, value in emergency_items:
            sheet[f'A{row}'] = label
            sheet[f'B{row}'] = value
            row += 1
        
        row += 1
        
        # Growth Fund
        sheet[f'A{row}'] = "Growth Investment Fund"
        sheet[f'A{row}'].font = Font(size=14, bold=True)
        row += 1
        
        growth_fund = recommendations.get('growthFund', {})
        growth_items = [
            ('Total Budget', f"${growth_fund.get('totalBudget', 0):,.2f}"),
            ('Marketing Budget', f"${growth_fund.get('marketingBudget', 0):,.2f}"),
            ('Hiring Budget', f"${growth_fund.get('hiringBudget', 0):,.2f}"),
            ('Equipment Upgrade', f"${growth_fund.get('equipmentUpgrade', 0):,.2f}"),
            ('R&D Budget', f"${growth_fund.get('researchDevelopment', 0):,.2f}")
        ]
        
        for label, value in growth_items:
            sheet[f'A{row}'] = label
            sheet[f'B{row}'] = value
            row += 1
        
        # Auto-adjust column widths
        for column in sheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            sheet.column_dimensions[column_letter].width = adjusted_width

